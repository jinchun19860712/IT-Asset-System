"""设备数据导入 / 导出（以「网络设备台账」Excel 为模板）

导入
- POST /devices/import-preview  上传 .xlsx，只做解析与差异分析，不写库（试运行）
- POST /devices/import          上传 .xlsx 正式导入（单事务，失败整体回滚）

导出
- GET  /devices/export          导出设备为 .xlsx，可选携带「端口」「机柜」Sheet
- GET  /devices/import-template 下载空白导入模板（含三个 Sheet 的表头与填写说明）

工作簿结构（导入/导出保持一致，可直接往返）
  Sheet1「设备台账」 台账主表（模板 29 列 + 文件夹/部门/机柜等扩展列）
  Sheet2「端口」     设备端口明细（上下联、聚合、堆叠、VLAN、对端）
  Sheet3「机柜」     机柜清单（U 数、位置、排号）
"""
import io
import re
from datetime import datetime, date
from typing import Optional
from urllib.parse import quote

from fastapi import APIRouter, Depends, UploadFile, File, HTTPException, Query, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.database import get_db
from app import models, schemas, crud

router = APIRouter(prefix="/devices", tags=["数据导入导出"])

SHEET_DEVICES = "设备台账"
SHEET_PORTS = "端口"
SHEET_RACKS = "机柜"

# ---------------------------------------------------------------- 表头定义

# 与原始台账模板一致的 29 列（顺序不能动，保证兼容用户手上的旧表）
TEMPLATE_HEADERS = [
    "No.", "Device Name", "Device type", "Device brand", "Parent Device Name",
    "Device model", "Area", "Room Type", "Room Number", "Service coding",
    "Asset Status", "Date[Disuse/Maintenance]",
    "Support SSH2", "Support Telnet", "Support Web", "Support SNMP",
    "Support RDP", "Support Console",
    "Management Service", "Management IP", "Network Mask", "Management VLAN",
    "Management MAC", "BMC IP", "BMC MAC",
    "User Name", "Password", "Function", "Remark",
]

# 扩展列（追加在模板列之后，旧表没有这些列也能正常导入）
EXTRA_HEADERS = [
    "Folder Path", "Department", "Supplier", "Asset Category",
    "Rack Name", "Rack Position(U)", "Rack Units", "Rack Face", "Port Count",
]

DEVICE_HEADERS = TEMPLATE_HEADERS + EXTRA_HEADERS

PORT_HEADERS = [
    "Device Name", "Port Name", "Port Type", "Connection Type",
    "Peer Device", "Peer Port", "LAG Group", "LAG Mode", "Stack ID",
    "VLAN", "Speed", "Description", "Sort",
]

RACK_HEADERS = [
    "Rack Name", "U Height", "Location", "Row Label", "Folder Path",
    "Description", "Mounted Devices",
]

# 表头（归一化小写）-> 设备字段；下划线开头的为需要二次解析的特殊字段
HEADER_MAP = {
    "device name": "name", "设备名称": "name", "设备名": "name", "name": "name",
    "device type": "device_type", "设备类型": "device_type", "type": "device_type",
    "device brand": "brand", "品牌": "brand", "brand": "brand",
    "parent device name": "_parent_name", "上级设备": "_parent_name", "父设备": "_parent_name",
    "device model": "model", "型号": "model", "model": "model",
    "area": "area", "区域": "area",
    "room type": "room_type", "房间类型": "room_type",
    "room number": "room_number", "房间号": "room_number",
    "service coding": "service_code", "资产编号": "service_code", "service code": "service_code",
    "asset status": "_status_raw", "资产状态": "_status_raw", "status": "_status_raw",
    "date[disuse/maintenance]": "disuse_date", "退运日期": "disuse_date", "维修日期": "disuse_date",
    "support ssh2": "support_ssh2",
    "support telnet": "support_telnet",
    "support web": "support_web",
    "support snmp": "support_snmp",
    "support rdp": "support_rdp",
    "support console": "support_console",
    "management service": "management_services", "管理服务": "management_services",
    "management ip": "ip_address", "管理ip": "ip_address", "ip": "ip_address",
    "network mask": "network_mask", "子网掩码": "network_mask", "掩码": "network_mask",
    "management vlan": "management_vlan", "管理vlan": "management_vlan",
    "management mac": "mac_address", "管理mac": "mac_address", "mac": "mac_address",
    "bmc ip": "bmc_ip",
    "bmc mac": "bmc_mac",
    "user name": "mgmt_username", "用户名": "mgmt_username", "username": "mgmt_username",
    "password": "mgmt_password", "密码": "mgmt_password",
    "function": "description", "功能": "description", "用途": "description",
    "remark": "remark", "备注": "remark",
    # 扩展列
    "folder path": "_folder_path", "文件夹": "_folder_path", "文件夹路径": "_folder_path",
    "department": "department", "部门": "department",
    "supplier": "supplier", "供应商": "supplier",
    "asset category": "_asset_path", "资产分类": "_asset_path",
    "rack name": "_rack_name", "机柜": "_rack_name", "机柜名称": "_rack_name",
    "rack position(u)": "_rack_position", "rack position": "_rack_position",
    "u位": "_rack_position", "起始u位": "_rack_position",
    "rack units": "_rack_units", "占用u数": "_rack_units", "u数": "_rack_units",
    "rack face": "_rack_face", "面向": "_rack_face", "正反面": "_rack_face",
}

BOOL_FIELDS = {"support_ssh2", "support_telnet", "support_web",
               "support_snmp", "support_rdp", "support_console"}

PORT_HEADER_MAP = {
    "device name": "_device_name", "设备名称": "_device_name", "设备名": "_device_name",
    "port name": "port_name", "端口名": "port_name", "端口名称": "port_name",
    "port type": "port_type", "端口类型": "port_type",
    "connection type": "connection_type", "连接类型": "connection_type",
    "peer device": "_peer_device_name", "对端设备": "_peer_device_name",
    "peer port": "peer_port_name", "对端端口": "peer_port_name",
    "lag group": "lag_group", "聚合组": "lag_group",
    "lag mode": "lag_mode", "聚合模式": "lag_mode",
    "stack id": "stack_id", "堆叠成员": "stack_id", "堆叠id": "stack_id",
    "vlan": "vlan_info", "vlan信息": "vlan_info",
    "speed": "port_speed", "速率": "port_speed",
    "description": "description", "描述": "description", "备注": "description",
    "sort": "sort_order", "排序": "sort_order",
}

RACK_HEADER_MAP = {
    "rack name": "name", "机柜名称": "name", "机柜": "name",
    "u height": "u_height", "总u数": "u_height", "u数": "u_height",
    "location": "location", "位置": "location", "机房": "location",
    "row label": "row_label", "排号": "row_label", "列号": "row_label",
    "folder path": "_folder_path", "文件夹": "_folder_path",
    "description": "description", "描述": "description", "备注": "description",
}

# 资产状态：英文 -> 中文（与 device_statuses 表对应）
STATUS_EN2CN = {
    "inuse": "在用", "in use": "在用", "using": "在用",
    "idle": "闲置", "spare": "闲置", "free": "闲置",
    "scrap": "淘汰", "discard": "淘汰", "retire": "淘汰", "retired": "淘汰",
    "maintenance": "维修", "repair": "维修", "broken": "维修", "fault": "维修",
    "disuse": "停用", "stop": "停用", "disable": "停用", "disabled": "停用",
}
# 导出时中文 -> 英文（保持模板风格）
STATUS_CN2EN = {"在用": "InUse", "闲置": "Idle", "淘汰": "Scrap",
                "维修": "Maintenance", "停用": "Disuse"}

PORT_TYPES = {"uplink", "downlink", "peer"}
CONN_TYPES = {"access", "trunk", "hybrid", "aggregate", "stack", "routed"}
IP_RE = re.compile(r"^\d{1,3}(\.\d{1,3}){3}$")


def _valid_ip(v: str) -> bool:
    """点分十进制且每段 0-255 才算合法（999.1.1.1 这类要能识别出来）。"""
    if not IP_RE.match(v):
        return False
    return all(0 <= int(x) <= 255 for x in v.split("."))

# 用于跳过「有格式没内容」的空行时，判断这行是否真的完全为空
MEANINGFUL_MIN_CHARS = 1


# ---------------------------------------------------------------- 通用工具

def _to_str(v) -> str:
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    return str(v).strip()


def _to_bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "1", "是", "yes", "y", "t", "√", "支持")


def _to_int(v) -> Optional[int]:
    s = _to_str(v)
    if not s:
        return None
    m = re.search(r"-?\d+", s)
    return int(m.group()) if m else None


def _norm(v) -> str:
    return str(v).strip().lower() if v is not None else ""


def _build_col_map(header_row, mapping: dict) -> dict:
    """列下标 -> 目标字段"""
    col_map = {}
    for idx, cell in enumerate(header_row):
        key = _norm(cell)
        if key in mapping:
            col_map[idx] = mapping[key]
    return col_map


def _find_header_row(ws, must_have: list, must_not_have: list = ()):
    """在前 30 行内寻找表头行。

    must_have 全部命中且 must_not_have 全部未命中才算匹配，
    用于区分「设备台账 / 端口 / 机柜」三张表——它们有同名列（如 Device Name、Rack Name）。
    """
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if i > 30:
            break
        cells = {_norm(c) for c in row if c is not None}
        if all(k in cells for k in must_have) and not any(k in cells for k in must_not_have):
            return row, i
    return None, None


def _row_is_blank(row) -> bool:
    for c in row:
        if c is None:
            continue
        if isinstance(c, str):
            if c.strip():
                return False
        else:
            return False
    return True


# ---------------------------------------------------------------- 解析

def parse_devices_sheet(wb) -> tuple:
    """解析设备主表。返回 (records, issues)。

    issues: [{row, level: error/warning, message}]
    """
    # 先找「既有 Device Name 又有 Device type」的标准台账表；
    # 找不到再放宽为「有 Device Name 且不是端口表」
    ws = header = header_no = None
    for must, forbid in ((["device name", "device type"], ["port name"]),
                         (["device name"], ["port name", "rack name"])):
        for cand in wb.worksheets:
            row, no = _find_header_row(cand, must, forbid)
            if row is not None:
                ws, header, header_no = cand, row, no
                break
        if ws is not None:
            break

    if header is None:
        raise ValueError("未找到设备表头（需包含 Device Name 列）")

    col_map = _build_col_map(header, HEADER_MAP)
    if "name" not in col_map.values():
        raise ValueError("表头缺少 Device Name 列")

    records, issues = [], []
    blank_rows = 0
    seen_names = {}

    rows = list(ws.iter_rows(values_only=True))
    for excel_row, row in enumerate(rows[header_no:], start=header_no + 1):
        if _row_is_blank(row):
            blank_rows += 1
            continue

        rec = {"_row": excel_row}
        blanks = set()      # 本行哪些列为空 —— 更新时据此跳过，避免把已有值清空
        for idx, field in col_map.items():
            raw = row[idx] if idx < len(row) else None
            if raw is None or (isinstance(raw, str) and not raw.strip()):
                blanks.add(field)
            if field in BOOL_FIELDS:
                rec[field] = _to_bool(raw)
            elif field in ("_rack_position", "_rack_units"):
                rec[field] = _to_int(raw)
            else:
                rec[field] = _to_str(raw)
        # 表里没有的列等同于空
        blanks |= (set(HEADER_MAP.values()) - set(col_map.values()))
        rec["_blank"] = blanks

        name = rec.get("name", "")
        if not name:
            # 有内容但没设备名 —— 属于需要提醒的脏行，而不是静默跳过
            issues.append({"row": excel_row, "level": "error",
                           "message": "该行有内容但缺少 Device Name，已跳过"})
            continue

        if name in seen_names:
            issues.append({"row": excel_row, "level": "warning",
                           "message": f"设备名「{name}」与第 {seen_names[name]} 行重复，后者覆盖前者"})
        seen_names[name] = excel_row

        # 父设备占位符统一成 None
        p = rec.get("_parent_name")
        if p in ("-", "", "—", "无", None):
            rec["_parent_name"] = None

        # 轻量字段校验（只提醒不拦截）
        for label, key in (("管理IP", "ip_address"), ("BMC IP", "bmc_ip")):
            v = rec.get(key)
            if v and v != "-" and not _valid_ip(v):
                issues.append({"row": excel_row, "level": "warning",
                               "message": f"{label}「{v}」格式可疑"})

        face = _norm(rec.get("_rack_face"))
        if face in ("背面", "后", "rear", "back"):
            rec["_rack_face"] = "rear"
        elif face in ("正面", "前", "front"):
            rec["_rack_face"] = "front"
        elif face:
            rec["_rack_face"] = "front"

        records.append(rec)

    return records, issues, blank_rows


def parse_ports_sheet(wb) -> tuple:
    """解析端口 Sheet，没有则返回空。"""
    for ws in wb.worksheets:
        header, header_no = _find_header_row(ws, ["port name"], ["device type"])
        if header is None:
            continue
        col_map = _build_col_map(header, PORT_HEADER_MAP)
        if "port_name" not in col_map.values():
            continue

        records, issues = [], []
        rows = list(ws.iter_rows(values_only=True))
        for excel_row, row in enumerate(rows[header_no:], start=header_no + 1):
            if _row_is_blank(row):
                continue
            rec = {"_row": excel_row}
            for idx, field in col_map.items():
                raw = row[idx] if idx < len(row) else None
                rec[field] = _to_int(raw) if field == "sort_order" else _to_str(raw)
            if not rec.get("_device_name") or not rec.get("port_name"):
                issues.append({"row": excel_row, "level": "error",
                               "message": "端口行缺少 Device Name 或 Port Name，已跳过",
                               "sheet": ws.title})
                continue
            pt = _norm(rec.get("port_type")) or "downlink"
            rec["port_type"] = pt if pt in PORT_TYPES else "downlink"
            ct = _norm(rec.get("connection_type")) or "access"
            rec["connection_type"] = ct if ct in CONN_TYPES else "access"
            records.append(rec)
        return records, issues
    return [], []


def parse_racks_sheet(wb) -> tuple:
    """解析机柜 Sheet，没有则返回空。"""
    # 设备台账里也有 Rack Name 列，必须排除掉，否则会把设备行当成机柜行
    for ws in wb.worksheets:
        header, header_no = _find_header_row(ws, ["rack name"],
                                             ["device name", "port name"])
        if header is None:
            continue
        col_map = _build_col_map(header, RACK_HEADER_MAP)
        if "name" not in col_map.values():
            continue

        records, issues = [], []
        rows = list(ws.iter_rows(values_only=True))
        for excel_row, row in enumerate(rows[header_no:], start=header_no + 1):
            if _row_is_blank(row):
                continue
            rec = {"_row": excel_row}
            for idx, field in col_map.items():
                raw = row[idx] if idx < len(row) else None
                rec[field] = _to_int(raw) if field == "u_height" else _to_str(raw)
            if not rec.get("name"):
                issues.append({"row": excel_row, "level": "error",
                               "message": "机柜行缺少 Rack Name，已跳过", "sheet": ws.title})
                continue
            rec["u_height"] = rec.get("u_height") or 42
            records.append(rec)
        return records, issues
    return [], []


# ---------------------------------------------------------------- 解析辅助

def _resolve_status(raw, status_name2id: dict, default_id):
    if not raw:
        return default_id
    s = str(raw).strip()
    low = s.lower()
    if low in STATUS_EN2CN:
        return status_name2id.get(STATUS_EN2CN[low], default_id)
    if s in status_name2id:
        return status_name2id[s]
    for cn, sid in status_name2id.items():
        if cn and cn in s:
            return sid
    return default_id


def _folder_path_index(db: Session) -> dict:
    """建立 {全路径(小写): folder_id} 索引，同时收录 - 与 / 两种分隔风格。"""
    index = {}
    for f in db.query(models.Folder).all():
        p = crud.get_folder_full_path(db, f.id, sep="-")
        if p:
            index[p.lower()] = f.id
            index[p.replace("-", "/").lower()] = f.id
        index.setdefault(f.name.lower(), f.id)
    return index


def _resolve_folder(db: Session, path_str: str, index: dict, auto_create: bool):
    """按全路径解析文件夹，找不到时可自动创建整条链。"""
    if not path_str:
        return None, None
    key = path_str.strip().lower()
    if key in index:
        return index[key], None

    sep = "/" if "/" in path_str else "-"
    parts = [p.strip() for p in path_str.split(sep) if p.strip()]
    if not parts:
        return None, None

    if not auto_create:
        return None, f"文件夹「{path_str}」不存在"

    parent_id = None
    parent_path = ""
    for depth, part in enumerate(parts):
        node = (db.query(models.Folder)
                .filter(models.Folder.name == part,
                        models.Folder.parent_id == parent_id)
                .first())
        if not node:
            node = models.Folder(name=part, parent_id=parent_id,
                                 is_department=(depth == 1))
            db.add(node)
            db.flush()          # 拿到 id，但不 commit —— 保证 dry_run 能整体回滚
            node.path = f"{parent_path}{node.id}/" if parent_path else f"/{node.id}/"
            db.flush()
        parent_id = node.id
        parent_path = node.path or f"/{node.id}/"
    index[key] = parent_id
    return parent_id, None


# ---------------------------------------------------------------- 导入核心

def _run_import(db: Session, data: bytes, *, mode: str, folder_id: Optional[int],
                import_ports: bool, import_racks: bool,
                auto_create_folder: bool, auto_create_rack: bool,
                port_mode: str, blank_policy: str, dry_run: bool) -> dict:
    """解析并（可选）写入。dry_run=True 时全部操作在事务内执行后回滚。"""
    try:
        wb = openpyxl.load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"无法读取 Excel：{e}")

    try:
        dev_records, issues, blank_rows = parse_devices_sheet(wb)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"解析失败：{e}")

    port_records, port_issues = parse_ports_sheet(wb) if import_ports else ([], [])
    rack_records, rack_issues = parse_racks_sheet(wb) if import_racks else ([], [])
    issues = issues + port_issues + rack_issues

    if not dev_records and not port_records and not rack_records:
        raise HTTPException(status_code=400,
                            detail="未解析到任何有效数据行（请检查表头与内容）")

    statuses = db.query(models.DeviceStatus).all()
    status_name2id = {s.name: s.id for s in statuses}
    default_status_id = status_name2id.get("在用") or (statuses[0].id if statuses else None)

    existing_devices = {d.name: d for d in db.query(models.Device).all()}
    existing_racks = {r.name: r for r in db.query(models.Rack).all()}
    folder_index = _folder_path_index(db)

    allowed = set(schemas.DeviceCreate.model_fields) - {"custom_values", "ports"}

    report = {
        "created": 0, "updated": 0, "skipped": 0, "unchanged": 0,
        "racks_created": 0, "racks_updated": 0,
        "mounted": 0, "ports_created": 0, "ports_updated": 0,
        "parents_linked": 0, "blank_rows": blank_rows,
        "rows": [], "issues": issues,
    }

    # ---- 阶段 1：机柜 ----
    for rec in rack_records:
        name = rec["name"]
        rk = existing_racks.get(name)
        payload = {k: v for k, v in rec.items()
                   if k in ("u_height", "location", "row_label", "description") and v != ""}
        fid, ferr = _resolve_folder(db, rec.get("_folder_path", ""), folder_index,
                                    auto_create_folder)
        if ferr:
            issues.append({"row": rec["_row"], "level": "warning",
                           "message": f"机柜「{name}」{ferr}", "sheet": SHEET_RACKS})
        if fid:
            payload["folder_id"] = fid

        if rk:
            for k, v in payload.items():
                setattr(rk, k, v)
            report["racks_updated"] += 1
        else:
            rk = models.Rack(name=name, **payload)
            db.add(rk)
            db.flush()
            existing_racks[name] = rk
            report["racks_created"] += 1

    # ---- 阶段 2：设备 upsert ----
    for rec in dev_records:
        name = rec["name"]
        row_no = rec["_row"]
        exists = existing_devices.get(name)

        if exists and mode == "insert_only":
            report["skipped"] += 1
            report["rows"].append({"row": row_no, "name": name, "action": "skip",
                                   "message": "已存在（仅新增模式）"})
            continue
        if not exists and mode == "update_only":
            report["skipped"] += 1
            report["rows"].append({"row": row_no, "name": name, "action": "skip",
                                   "message": "不存在（仅更新模式）"})
            continue

        blanks = rec.get("_blank", set())
        keep_blank = (blank_policy == "ignore") and bool(exists)

        dev_data = {k: v for k, v in rec.items() if k in allowed}
        if keep_blank:
            # 更新已有设备时，Excel 里留空的列不参与写入（部分列更新场景不误清空）
            dev_data = {k: v for k, v in dev_data.items() if k not in blanks}
        if not (keep_blank and "_status_raw" in blanks):
            dev_data["status_id"] = _resolve_status(rec.get("_status_raw"),
                                                    status_name2id, default_status_id)

        # 文件夹：行内 Folder Path 优先，其次接口传入的 folder_id
        fid, ferr = _resolve_folder(db, rec.get("_folder_path", ""), folder_index,
                                    auto_create_folder)
        if ferr:
            issues.append({"row": row_no, "level": "warning", "message": ferr})
        if fid:
            dev_data["folder_id"] = fid
        elif folder_id and not exists:
            dev_data["folder_id"] = folder_id

        if not dev_data.get("department") and dev_data.get("folder_id"):
            dev_data["department"] = crud.get_folder_department(db, dev_data["folder_id"])

        # 资产分类：只做匹配，不自动建节点，避免污染「设备资产」树
        asset_path = rec.get("_asset_path", "")
        if asset_path:
            aid, aerr = _resolve_folder(db, asset_path, folder_index, False)
            if aid:
                dev_data["asset_folder_id"] = aid
            else:
                issues.append({"row": row_no, "level": "warning",
                               "message": f"资产分类「{asset_path}」不存在，已跳过该列"})

        try:
            schemas.DeviceCreate(**dev_data)      # 校验
        except Exception as e:
            report["skipped"] += 1
            issues.append({"row": row_no, "level": "error",
                           "message": f"「{name}」字段非法：{e}"})
            report["rows"].append({"row": row_no, "name": name, "action": "error",
                                   "message": str(e)[:160]})
            continue

        if exists:
            changes = {}
            for k, v in dev_data.items():
                old = getattr(exists, k, None)
                if (old or "") != (v or ""):
                    changes[k] = [_to_str(old), _to_str(v)]
                    setattr(exists, k, v)
            if changes:
                report["updated"] += 1
                report["rows"].append({"row": row_no, "name": name, "action": "update",
                                       "changes": changes})
            else:
                report["unchanged"] += 1
                report["rows"].append({"row": row_no, "name": name, "action": "unchanged"})
            obj = exists
        else:
            obj = models.Device(**dev_data)
            db.add(obj)
            db.flush()
            existing_devices[name] = obj
            report["created"] += 1
            report["rows"].append({"row": row_no, "name": name, "action": "create"})

        rec["_dev_id"] = obj.id

    # ---- 阶段 3：父子关系（第二遍，支持文件内互相引用）----
    for rec in dev_records:
        did, pname = rec.get("_dev_id"), rec.get("_parent_name")
        if not did or not pname:
            continue
        parent = existing_devices.get(pname)
        if not parent:
            issues.append({"row": rec["_row"], "level": "warning",
                           "message": f"父设备「{pname}」不存在（可能是虚拟堆叠名），已忽略父子关系"})
            continue
        if parent.id == did:
            issues.append({"row": rec["_row"], "level": "warning",
                           "message": f"「{rec['name']}」父设备指向自身，已忽略"})
            continue
        dev = db.get(models.Device, did)
        if dev and dev.parent_device_id != parent.id:
            dev.parent_device_id = parent.id
            report["parents_linked"] += 1
    db.flush()

    # ---- 阶段 4：机柜上架 ----
    if import_racks:
        for rec in dev_records:
            did = rec.get("_dev_id")
            rname = rec.get("_rack_name")
            if not did or not rname:
                continue
            rk = existing_racks.get(rname)
            if not rk:
                if not auto_create_rack:
                    issues.append({"row": rec["_row"], "level": "warning",
                                   "message": f"机柜「{rname}」不存在，未上架"})
                    continue
                rk = models.Rack(name=rname, u_height=42)
                db.add(rk)
                db.flush()
                existing_racks[rname] = rk
                report["racks_created"] += 1

            pos = rec.get("_rack_position")
            units = rec.get("_rack_units") or 1
            face = rec.get("_rack_face") or "front"
            if not pos:
                issues.append({"row": rec["_row"], "level": "warning",
                               "message": f"「{rec['name']}」指定了机柜但缺少 U 位，未上架"})
                continue
            conflict = crud.check_rack_conflict(db, rk.id, pos, units, face,
                                                exclude_device_id=did)
            if conflict:
                issues.append({"row": rec["_row"], "level": "warning",
                               "message": f"「{rec['name']}」上架失败：{conflict}"})
                continue
            dev = db.get(models.Device, did)
            dev.rack_id, dev.rack_position = rk.id, pos
            dev.rack_units, dev.rack_face = units, face
            # session 关闭了 autoflush，必须显式 flush，
            # 否则同一次导入中的后续设备查不到刚上架的这台，U位冲突检测会漏判
            db.flush()
            report["mounted"] += 1

    # ---- 阶段 5：端口 ----
    if import_ports and port_records:
        touched = set()
        if port_mode == "replace":
            for rec in port_records:
                dev = existing_devices.get(rec["_device_name"])
                if dev and dev.id not in touched:
                    db.query(models.DevicePort).filter(
                        models.DevicePort.device_id == dev.id).delete()
                    touched.add(dev.id)

        for order, rec in enumerate(port_records):
            dev = existing_devices.get(rec["_device_name"])
            if not dev:
                issues.append({"row": rec["_row"], "level": "warning",
                               "message": f"端口所属设备「{rec['_device_name']}」不存在，已跳过",
                               "sheet": SHEET_PORTS})
                continue
            peer = existing_devices.get(rec.get("_peer_device_name") or "")
            if rec.get("_peer_device_name") and not peer:
                issues.append({"row": rec["_row"], "level": "warning",
                               "message": f"对端设备「{rec['_peer_device_name']}」不存在，仅保留端口",
                               "sheet": SHEET_PORTS})

            fields = {
                "port_type": rec.get("port_type") or "downlink",
                "connection_type": rec.get("connection_type") or "access",
                "peer_device_id": peer.id if peer else None,
                "peer_port_name": rec.get("peer_port_name") or "",
                "lag_group": rec.get("lag_group") or "",
                "lag_mode": rec.get("lag_mode") or "",
                "stack_id": rec.get("stack_id") or "",
                "vlan_info": rec.get("vlan_info") or "",
                "port_speed": rec.get("port_speed") or "",
                "description": rec.get("description") or "",
                "sort_order": rec.get("sort_order") if rec.get("sort_order") is not None else order,
            }

            port = None
            if port_mode != "replace":
                port = (db.query(models.DevicePort)
                        .filter(models.DevicePort.device_id == dev.id,
                                models.DevicePort.port_name == rec["port_name"])
                        .first())
            if port:
                for k, v in fields.items():
                    setattr(port, k, v)
                report["ports_updated"] += 1
            else:
                db.add(models.DevicePort(device_id=dev.id,
                                         port_name=rec["port_name"], **fields))
                report["ports_created"] += 1

    report["issues"] = issues
    report["error_count"] = sum(1 for i in issues if i.get("level") == "error")
    report["warning_count"] = sum(1 for i in issues if i.get("level") == "warning")

    if dry_run:
        db.rollback()
    else:
        db.commit()
    return report


def _summary_text(r: dict, dry: bool) -> str:
    head = "预览完成" if dry else "导入完成"
    parts = [f"新增 {r['created']}", f"更新 {r['updated']}", f"无变化 {r['unchanged']}",
             f"跳过 {r['skipped']}"]
    if r["racks_created"] or r["racks_updated"]:
        parts.append(f"机柜 +{r['racks_created']}/~{r['racks_updated']}")
    if r["mounted"]:
        parts.append(f"上架 {r['mounted']}")
    if r["ports_created"] or r["ports_updated"]:
        parts.append(f"端口 +{r['ports_created']}/~{r['ports_updated']}")
    if r["parents_linked"]:
        parts.append(f"父子关系 {r['parents_linked']}")
    tail = f"；问题 {r['error_count']} 错误 / {r['warning_count']} 警告" \
        if (r["error_count"] or r["warning_count"]) else ""
    return f"{head}：" + "，".join(parts) + tail


# ---------------------------------------------------------------- 导入接口

@router.post("/import-preview")
async def preview_import(
    file: UploadFile = File(...),
    mode: str = Form("upsert"),
    folder_id: Optional[int] = Form(None),
    import_ports: bool = Form(True),
    import_racks: bool = Form(True),
    auto_create_folder: bool = Form(False),
    auto_create_rack: bool = Form(False),
    port_mode: str = Form("merge"),
    blank_policy: str = Form("ignore"),
    db: Session = Depends(get_db),
):
    """试运行：解析并计算差异，但不写库（事务内执行后回滚）。"""
    _check_file(file)
    data = await file.read()
    try:
        report = _run_import(db, data, mode=mode, folder_id=folder_id,
                             import_ports=import_ports, import_racks=import_racks,
                             auto_create_folder=auto_create_folder,
                             auto_create_rack=auto_create_rack,
                             port_mode=port_mode, blank_policy=blank_policy,
                             dry_run=True)
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=f"预览失败：{e}")
    return {"code": 0, "message": _summary_text(report, True), "data": report}


@router.post("/import")
async def import_devices(
    file: UploadFile = File(...),
    mode: str = Form("upsert"),
    folder_id: Optional[int] = Form(None),
    import_ports: bool = Form(True),
    import_racks: bool = Form(True),
    auto_create_folder: bool = Form(False),
    auto_create_rack: bool = Form(False),
    port_mode: str = Form("merge"),
    blank_policy: str = Form("ignore"),
    db: Session = Depends(get_db),
):
    """正式导入。整个过程在单个事务中完成，任何未捕获异常都会整体回滚。

    mode: upsert(默认) / insert_only(仅新增) / update_only(仅更新)
    port_mode: merge(按端口名增量更新，默认) / replace(整机替换端口)
    """
    _check_file(file)
    if mode not in ("upsert", "insert_only", "update_only"):
        raise HTTPException(status_code=400, detail="mode 仅支持 upsert / insert_only / update_only")
    if port_mode not in ("merge", "replace"):
        raise HTTPException(status_code=400, detail="port_mode 仅支持 merge / replace")
    if blank_policy not in ("ignore", "overwrite"):
        raise HTTPException(status_code=400, detail="blank_policy 仅支持 ignore / overwrite")

    data = await file.read()
    try:
        report = _run_import(db, data, mode=mode, folder_id=folder_id,
                             import_ports=import_ports, import_racks=import_racks,
                             auto_create_folder=auto_create_folder,
                             auto_create_rack=auto_create_rack,
                             port_mode=port_mode, blank_policy=blank_policy,
                             dry_run=False)
    except HTTPException:
        db.rollback()
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败，已回滚：{e}")
    return {"code": 0, "message": _summary_text(report, False), "data": report}


def _check_file(file: UploadFile):
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx / .xlsm 文件")


# ---------------------------------------------------------------- 导出

@router.get("/export")
def export_devices(
    include_ports: bool = Query(True, description="附带「端口」Sheet"),
    include_racks: bool = Query(True, description="附带「机柜」Sheet"),
    folder_id: Optional[int] = Query(None, description="组织机构范围：只导出该文件夹及其子孙"),
    asset_folder_id: Optional[int] = Query(None, description="设备资产范围：只导出该资产分类及其子孙"),
    device_types: Optional[str] = Query(None, description="设备类型筛选，多个用逗号分隔"),
    db: Session = Depends(get_db),
):
    """导出设备为 Excel（列与导入模板一致，可直接改完再导回来）。"""
    query = (db.query(models.Device)
             .options(joinedload(models.Device.status),
                      joinedload(models.Device.parent_device),
                      joinedload(models.Device.rack),
                      joinedload(models.Device.folder),
                      joinedload(models.Device.ports).joinedload(models.DevicePort.peer_device)))
    if folder_id:
        ids = crud.get_folder_descendants(db, folder_id)
        if ids:
            query = query.filter(models.Device.folder_id.in_(ids))
    if asset_folder_id:
        ids = crud.get_folder_descendants(db, asset_folder_id)
        if ids:
            query = query.filter(models.Device.asset_folder_id.in_(ids))
    if device_types:
        types = [t.strip() for t in device_types.split(",") if t.strip()]
        if types:
            query = query.filter(models.Device.device_type.in_(types))
    devices = query.order_by(models.Device.id).all()

    wb = openpyxl.Workbook()

    # ---- Sheet1 设备台账 ----
    ws = wb.active
    ws.title = SHEET_DEVICES
    ws.append(DEVICE_HEADERS)
    path_cache = {}
    asset_path_cache = {}
    for i, d in enumerate(devices, start=1):
        status_name = d.status.name if d.status else ""
        if d.folder_id not in path_cache:
            path_cache[d.folder_id] = crud.get_folder_full_path(db, d.folder_id) if d.folder_id else ""
        ws.append([
            i,
            d.name, d.device_type, d.brand,
            d.parent_device.name if d.parent_device else "-",
            d.model, d.area, d.room_type, d.room_number, d.service_code,
            STATUS_CN2EN.get(status_name, status_name),
            d.disuse_date,
            d.support_ssh2, d.support_telnet, d.support_web,
            d.support_snmp, d.support_rdp, d.support_console,
            d.management_services,
            d.ip_address, d.network_mask, d.management_vlan, d.mac_address,
            d.bmc_ip, d.bmc_mac,
            d.mgmt_username, d.mgmt_password,
            d.description, d.remark or "",
            # 扩展列
            path_cache[d.folder_id], d.department or "",
            d.supplier or "",
            asset_path_cache.setdefault(
                d.asset_folder_id,
                crud.get_folder_full_path(db, d.asset_folder_id) if d.asset_folder_id else ""
            ),
            d.rack.name if d.rack else "",
            d.rack_position or "", d.rack_units or "",
            d.rack_face or "", len(d.ports or []),
        ])
    _style_sheet(ws, len(DEVICE_HEADERS))

    # ---- Sheet2 端口 ----
    if include_ports:
        wp = wb.create_sheet(SHEET_PORTS)
        wp.append(PORT_HEADERS)
        for d in devices:
            for p in sorted(d.ports or [], key=lambda x: (x.sort_order or 0, x.id)):
                wp.append([
                    d.name, p.port_name, p.port_type, p.connection_type,
                    p.peer_device.name if p.peer_device else "",
                    p.peer_port_name or "", p.lag_group or "", p.lag_mode or "",
                    p.stack_id or "", p.vlan_info or "", p.port_speed or "",
                    p.description or "", p.sort_order or 0,
                ])
        _style_sheet(wp, len(PORT_HEADERS))

    # ---- Sheet3 机柜 ----
    if include_racks:
        wr = wb.create_sheet(SHEET_RACKS)
        wr.append(RACK_HEADERS)
        racks = db.query(models.Rack).order_by(models.Rack.id).all()
        for r in racks:
            mounted = [x for x in (r.devices or []) if x.rack_position]
            wr.append([
                r.name, r.u_height, r.location or "", r.row_label or "",
                crud.get_folder_full_path(db, r.folder_id) if r.folder_id else "",
                r.description or "", len(mounted),
            ])
        _style_sheet(wr, len(RACK_HEADERS))

    return _xlsx_response(wb, f"网络设备台账-{datetime.now():%Y%m%d}.xlsx")


@router.get("/import-template")
def download_template(with_sample: bool = Query(False, description="附带一行示例数据")):
    """下载空白导入模板（三个 Sheet + 填写说明）。"""
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = SHEET_DEVICES
    ws.append(DEVICE_HEADERS)
    if with_sample:
        ws.append([
            1, "sw-h3c-s5130-01", "Switch", "H3C", "rt-h3c-msr5660", "S5130S-28S-EI",
            "行政楼", "网络机房", "301", "ZC-2024-001", "InUse", "",
            True, True, True, True, False, True,
            "SSH2,Web,SNMP", "172.20.10.11", "255.255.255.0", "VLAN 7",
            "00-11-22-33-44-55", "", "", "admin", "Admin@123",
            "接入交换机", "季度巡检",
            "黄山健康职业学院/信息中心/网络设备", "信息中心",
            "A-01", 10, 1, "front", 2,
        ])
    _style_sheet(ws, len(DEVICE_HEADERS))

    wp = wb.create_sheet(SHEET_PORTS)
    wp.append(PORT_HEADERS)
    if with_sample:
        wp.append(["sw-h3c-s5130-01", "GigabitEthernet1/0/1", "uplink", "aggregate",
                   "rt-h3c-msr5660", "GE0/0/1", "BAGG1", "lacp", "", "", "1G",
                   "上联出口路由器", 1])
        wp.append(["sw-h3c-s5130-01", "GigabitEthernet1/0/24", "downlink", "access",
                   "", "", "", "", "", "VLAN 20", "1G", "接入办公区", 2])
    _style_sheet(wp, len(PORT_HEADERS))

    wr = wb.create_sheet(SHEET_RACKS)
    wr.append(RACK_HEADERS)
    if with_sample:
        wr.append(["A-01", 42, "行政楼三楼网络机房", "A 排",
                   "黄山健康职业学院/信息中心", "核心机柜", 1])
    _style_sheet(wr, len(RACK_HEADERS))

    _append_readme(wb)
    return _xlsx_response(wb, "网络设备台账-导入模板.xlsx")


def _append_readme(wb):
    ws = wb.create_sheet("填写说明")
    lines = [
        ["字段", "说明"],
        ["Device Name", "必填，设备唯一标识。导入时按此列判断新增还是更新"],
        ["Parent Device Name", "填上级设备的 Device Name；没有填 - 。可引用同一份文件里的设备"],
        ["Asset Status", "InUse / Idle / Scrap / Maintenance / Disuse，也可直接填中文"],
        ["Support *", "TRUE / FALSE（也接受 是 / 1 / Y）"],
        ["Folder Path", "文件夹全路径，支持 A/B/C 或 A-B-C；开启「自动创建文件夹」时可新建"],
        ["Department", "留空则按 Folder Path 自动推导所属部门"],
        ["Supplier", "供应商名称，与「基础数据 → 供应商」共用一份清单"],
        ["Asset Category", "设备资产树的全路径，如 设备资产-资产-IT资产；只做匹配，不会自动新建分类"],
        ["Rack Name / Rack Position(U)", "同时填写才会上架；U 位冲突会给出警告并跳过上架"],
        ["Rack Units / Rack Face", "占用 U 数（默认 1）、front 或 rear（默认 front）"],
        ["", ""],
        ["Sheet「端口」", "Device Name + Port Name 为定位键；Peer Device 填对端设备名"],
        ["Port Type", "uplink 上联 / downlink 下联 / peer 对等"],
        ["Connection Type", "access / trunk / hybrid / aggregate 聚合 / stack 堆叠 / routed"],
        ["LAG Group / Stack ID", "聚合组号（如 BAGG1）/ 堆叠成员号，按连接类型填写"],
        ["", ""],
        ["Sheet「机柜」", "Rack Name 为唯一键，U Height 默认 42"],
        ["", ""],
        ["导入模式", "upsert 有则更新无则新增（默认）/ insert_only 仅新增 / update_only 仅更新"],
        ["端口模式", "merge 按端口名增量更新（默认）/ replace 先清空该设备端口再写入"],
        ["空值策略", "ignore 空单元格不覆盖已有值（默认，可只填想改的列）/ overwrite 空单元格清空该字段"],
        ["建议流程", "先「导入预览」查看差异，确认无误后再正式导入"],
    ]
    for r in lines:
        ws.append(r)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 90
    for c in ws["A"]:
        c.font = Font(bold=True)
    ws["A1"].fill = PatternFill("solid", fgColor="DDEBF7")
    ws["B1"].fill = PatternFill("solid", fgColor="DDEBF7")


# ---------------------------------------------------------------- 输出辅助

def _style_sheet(ws, col_count: int):
    """表头加粗 + 底色 + 冻结首行 + 自适应列宽。"""
    fill = PatternFill("solid", fgColor="DDEBF7")
    for idx in range(1, col_count + 1):
        cell = ws.cell(row=1, column=idx)
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"

    for idx in range(1, col_count + 1):
        letter = get_column_letter(idx)
        longest = len(str(ws.cell(row=1, column=idx).value or ""))
        for row in range(2, min(ws.max_row, 200) + 1):
            v = ws.cell(row=row, column=idx).value
            if v is not None:
                longest = max(longest, len(str(v)))
        ws.column_dimensions[letter].width = min(max(longest + 3, 8), 34)


def _xlsx_response(wb, filename: str) -> StreamingResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": _content_disposition(filename)},
    )


def _content_disposition(filename: str) -> str:
    """生成兼容中文文件名的 Content-Disposition（RFC 5987）。"""
    ascii_name = filename.encode("ascii", "ignore").decode() or "download.xlsx"
    return f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{quote(filename)}"
